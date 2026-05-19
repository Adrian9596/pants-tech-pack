from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from .utils import (
    fill, bold_dark, bold_white, normal, input_font, thin_border, center, left, right_a,
    money_fmt, pct_fmt, merge_title, style_header, style_input, set_col_widths,
    disable_gridlines, freeze, set_print_area,
    NAVY, MED_BLUE, WHITE, ALT_ROW, GRAY_SUB, INPUT_BG, EXW_BG, FOB_BG,
    GREEN_OK, YELLOW_OK, RED_WARN
)
from openpyxl.utils import get_column_letter


def generate_costing(wb, req):
    ws = wb.create_sheet("Costing Sheet")
    disable_gridlines(ws)
    set_col_widths(ws, {1:3, 2:6, 3:38, 4:18, 5:14, 6:5})
    ws.row_dimensions[1].height = 8

    # ── Header block ──────────────────────────────────────────────
    _block_header(ws, 2, f"COSTING SHEET — {req.style_name}  |  {req.resolved_style_number()}", 16)
    meta = [
        ("Season", req.season or "TBD"),
        ("Currency", "USD"),
        ("Date", __import__('datetime').datetime.now().strftime("%Y-%m-%d")),
        ("Target FOB", req.target_fob),
    ]
    for i, (k, v) in enumerate(meta):
        r = 3 + i
        ws.row_dimensions[r].height = 16
        kc = ws.cell(row=r, column=2, value=k)
        kc.font = bold_dark(); kc.alignment = left; kc.fill = fill(ALT_ROW if i%2 else WHITE); kc.border = thin_border()
        vc = ws.cell(row=r, column=3, value=v)
        vc.font = normal(); vc.alignment = left; vc.fill = fill(ALT_ROW if i%2 else WHITE); vc.border = thin_border()
        for col in [4, 5]:
            _blank(ws, r, col, ALT_ROW if i%2 else WHITE)

    cur_row = 8

    # ── Section A: Material Costs ──────────────────────────────────
    cur_row = _section_title(ws, cur_row, "A", "MATERIAL COSTS (Linked from BOM)")
    has_bom = "BOM" in wb.sheetnames
    if has_bom:
        f_fabric = "=IFERROR(SUMIF('BOM'!B:B,\"Fabric\",'BOM'!M:M),0)"
        f_trim   = "=IFERROR(SUMIF('BOM'!B:B,\"Trim\",'BOM'!M:M),0)"
        f_label  = "=IFERROR(SUMIF('BOM'!B:B,\"Label\",'BOM'!M:M)+SUMIF('BOM'!B:B,\"Packaging\",'BOM'!M:M),0)"
    else:
        f_fabric = None
        f_trim   = None
        f_label  = None
    a_labels = [
        ("A1", "Total Fabric Cost",              f_fabric),
        ("A2", "Total Trim Cost",                f_trim),
        ("A3", "Total Label & Packaging Cost",   f_label),
        ("A4", "Total Material Cost",            "=A1_CELL+A2_CELL+A3_CELL"),
        ("A5", "Material Waste %",               req.costing.material_waste_pct, True),
        ("A6", "Material Cost incl. Waste",      "=A4_CELL*(1+A5_CELL)"),
    ]
    a_cells = {}
    for code, label, val, *args in a_labels:
        is_input = args[0] if args else False
        r = cur_row
        ws.row_dimensions[r].height = 16
        a_cells[code] = r
        _label_row(ws, r, code, label, val, is_input, bold=code in ("A4", "A6"))
        if code in ("A4", "A6"):
            _make_bold_row(ws, r)
        cur_row += 1

    # Fix formula references using actual row numbers
    for code in ("A4", "A6"):
        c = ws.cell(row=a_cells[code], column=4)
        if code == "A4":
            c.value = f"=D{a_cells['A1']}+D{a_cells['A2']}+D{a_cells['A3']}"
        else:
            c.value = f"=D{a_cells['A4']}*(1+D{a_cells['A5']})"
        money_fmt(c)

    cur_row += 1

    # ── Section B: CMT ────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "B", "CMT — CUT, MAKE, TRIM")
    co = req.costing
    b_labels = [
        ("B1", "Cut Cost",               co.cut_cost,         True),
        ("B2", "Make / Sewing Cost",     co.make_cost,        True),
        ("B3", "Trim / Finishing Labor", co.trim_finishing,   True),
        ("B4", "Total CMT",              None),
    ]
    b_cells = {}
    for code, label, val, *args in b_labels:
        is_input = args[0] if args else False
        r = cur_row; b_cells[code] = r
        ws.row_dimensions[r].height = 16
        _label_row(ws, r, code, label, val, is_input, bold=code == "B4")
        if code == "B4":
            ws.cell(row=r, column=4).value = f"=D{b_cells['B1']}+D{b_cells['B2']}+D{b_cells['B3']}"
            _make_bold_row(ws, r)
        cur_row += 1
    cur_row += 1

    # ── Section C: Other Direct ────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "C", "OTHER DIRECT COSTS")
    c_labels = [
        ("C1", "Washing / Treatment",      co.washing_treatment,   True),
        ("C2", "Embroidery / Print",       co.embroidery_print,    True),
        ("C3", "Hang Tag & Ticketing",     co.hang_tag_ticketing,  True),
        ("C4", "Freight (factory→port)",   co.freight_factory,     True),
        ("C5", "Testing / Compliance",     co.testing_compliance,  True),
        ("C6", "Total Other Direct",       None),
    ]
    c_cells = {}
    for code, label, val, *args in c_labels:
        is_input = args[0] if args else False
        r = cur_row; c_cells[code] = r
        ws.row_dimensions[r].height = 16
        _label_row(ws, r, code, label, val, is_input, bold=code == "C6")
        if code == "C6":
            ws.cell(row=r, column=4).value = (
                f"=D{c_cells['C1']}+D{c_cells['C2']}+D{c_cells['C3']}"
                f"+D{c_cells['C4']}+D{c_cells['C5']}"
            )
            _make_bold_row(ws, r)
        cur_row += 1
    cur_row += 1

    # ── Section D: Factory Buildup ─────────────────────────────────
    cur_row = _section_title(ws, cur_row, "D", "FACTORY COST BUILDUP")
    d_cells = {}
    d_spec = [
        ("D1", "Total Direct Cost",          None,                    False, False),
        ("D2", "Factory Overhead %",         co.overhead_pct,         True,  False),
        ("D3", "Factory Overhead Amount",    None,                    False, False),
        ("D4", "Factory Cost (ex-profit)",   None,                    False, True),
        ("D5", "Factory Profit Margin %",    co.margin_pct,           True,  False),
        ("D6", "Factory Profit Amount",      None,                    False, False),
        ("D7", "Ex-Factory Cost (EXW)",      None,                    False, True),
    ]
    for code, label, val, is_input, is_bold in d_spec:
        r = cur_row; d_cells[code] = r
        ws.row_dimensions[r].height = 16
        _label_row(ws, r, code, label, val, is_input, bold=is_bold)
        cur_row += 1

    # Wire D formulas
    _set_formula(ws, d_cells["D1"], f"=D{a_cells['A6']}+D{b_cells['B4']}+D{c_cells['C6']}", bold=False)
    _set_formula(ws, d_cells["D3"], f"=D{d_cells['D1']}*D{d_cells['D2']}", pct=False)
    _set_formula(ws, d_cells["D4"], f"=D{d_cells['D1']}+D{d_cells['D3']}", bold=True)
    _set_formula(ws, d_cells["D6"], f"=D{d_cells['D4']}*D{d_cells['D5']}", pct=False)
    _set_formula(ws, d_cells["D7"], f"=D{d_cells['D4']}+D{d_cells['D6']}", bold=True)

    # EXW background highlight
    for col in range(2, 6):
        ws.cell(row=d_cells["D7"], column=col).fill = fill(EXW_BG)
    cur_row += 1

    # ── Section E: FOB ────────────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "E", "FOB PRICE BUILDUP")
    e_cells = {}
    e_spec = [
        ("E1", "Export Documentation",   co.export_docs,      True),
        ("E2", "Local Transport to Port", co.local_transport,  True),
        ("E3", "FOB Price",               None),
    ]
    for code, label, val, *args in e_spec:
        is_input = args[0] if args else False
        r = cur_row; e_cells[code] = r
        ws.row_dimensions[r].height = 20 if code == "E3" else 16
        _label_row(ws, r, code, label, val, is_input, bold=code == "E3")
        cur_row += 1

    _set_formula(ws, e_cells["E3"], f"=D{d_cells['D7']}+D{e_cells['E1']}+D{e_cells['E2']}", bold=True)
    for col in range(2, 6):
        c = ws.cell(row=e_cells["E3"], column=col)
        c.fill = fill(FOB_BG)
        c.font = Font(name="Calibri", size=13, bold=True)
    cur_row += 1

    # ── Section F: Brand Margin ────────────────────────────────────
    cur_row = _section_title(ws, cur_row, "F", "BRAND MARGIN ANALYSIS")
    f_cells = {}
    f_spec = [
        ("F1", "Target Wholesale Price", co.target_wholesale, True),
        ("F2", "Target Retail Price",    co.target_retail,    True),
        ("F3", "Gross Margin (Wholesale)", None),
        ("F4", "Gross Margin (Retail)",    None),
        ("F5", "Keystone Multiple (Retail/FOB)", None),
        ("F6", "Markup on Cost (Retail/FOB)",    None),
    ]
    for code, label, val, *args in f_spec:
        is_input = args[0] if args else False
        r = cur_row; f_cells[code] = r
        ws.row_dimensions[r].height = 16
        _label_row(ws, r, code, label, val, is_input)
        cur_row += 1

    fob_r  = e_cells["E3"]
    _set_formula(ws, f_cells["F3"], f"=(D{f_cells['F1']}-D{fob_r})/D{f_cells['F1']}", pct=True)
    _set_formula(ws, f_cells["F4"], f"=(D{f_cells['F2']}-D{fob_r})/D{f_cells['F2']}", pct=True)
    _set_formula(ws, f_cells["F5"], f"=D{f_cells['F2']}/D{fob_r}", pct=False, fmt='0.0"x"')
    _set_formula(ws, f_cells["F6"], f"=(D{f_cells['F2']}-D{fob_r})/D{fob_r}", pct=True)

    # Traffic-light conditional formatting on F3/F4
    for f_row in (f_cells["F3"], f_cells["F4"]):
        cell_ref = f"D{f_row}"
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.55"],
                       fill=PatternFill("solid", fgColor=GREEN_OK))
        )
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="between", formula=["0.3", "0.5499"],
                       fill=PatternFill("solid", fgColor=YELLOW_OK))
        )
        ws.conditional_formatting.add(
            cell_ref,
            CellIsRule(operator="lessThan", formula=["0.3"],
                       fill=PatternFill("solid", fgColor=RED_WARN))
        )
    cur_row += 1

    # ── Section G: Cost Breakdown Summary ─────────────────────────
    cur_row = _section_title(ws, cur_row, "G", "COST BREAKDOWN SUMMARY")
    g_headers = ["Component", "Cost (USD)", "% of FOB"]
    hrow = cur_row
    ws.row_dimensions[hrow].height = 18
    for col, hdr in zip([2, 4, 5], g_headers):
        c = ws.cell(row=hrow, column=col, value=hdr)
        c.fill = fill(NAVY); c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.alignment = center; c.border = thin_border()
    cur_row += 1

    fob_ref = f"D{fob_r}"
    g_rows = [
        ("Fabric",          f"=D{a_cells['A1']}"),
        ("Trims",           f"=D{a_cells['A2']}"),
        ("Labels & Pkg",    f"=D{a_cells['A3']}"),
        ("CMT",             f"=D{b_cells['B4']}"),
        ("Other Direct",    f"=D{c_cells['C6']}"),
        ("Overhead",        f"=D{d_cells['D3']}"),
        ("Factory Profit",  f"=D{d_cells['D6']}"),
        ("FOB Add-ons",     f"=D{e_cells['E1']}+D{e_cells['E2']}"),
    ]
    for i, (name, cost_f) in enumerate(g_rows):
        r = cur_row
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        nc = ws.cell(row=r, column=2, value=name)
        nc.fill = fill(ALT_ROW if alt else WHITE); nc.font = normal(); nc.alignment = left; nc.border = thin_border()
        cc = ws.cell(row=r, column=4, value=cost_f)
        cc.fill = fill(ALT_ROW if alt else WHITE); cc.font = normal(); cc.alignment = right_a; cc.border = thin_border(); cc.number_format = "$#,##0.00"
        pc = ws.cell(row=r, column=5, value=f"=IF({fob_ref}=0,\"\",D{r}/{fob_ref})")
        pc.fill = fill(ALT_ROW if alt else WHITE); pc.font = normal(); pc.alignment = right_a; pc.border = thin_border(); pc.number_format = "0.0%"
        cur_row += 1

    # Total FOB row
    ws.row_dimensions[cur_row].height = 18
    for col, val, fmt in [(2, "TOTAL FOB", None), (4, f"={fob_ref}", "$#,##0.00"), (5, "100%", "0.0%")]:
        c = ws.cell(row=cur_row, column=col, value=val)
        c.fill = fill(FOB_BG); c.font = Font(name="Calibri", bold=True, size=11)
        c.alignment = center; c.border = thin_border()
        if fmt: c.number_format = fmt

    freeze(ws, "A6")
    set_print_area(ws, f"A1:F{cur_row + 3}")


# ── Helpers ───────────────────────────────────────────────────────

def _block_header(ws, row, text, size=14):
    ws.row_dimensions[row].height = 28
    from openpyxl.utils import get_column_letter
    ws.merge_cells(f"B{row}:E{row}")
    c = ws[f"B{row}"]
    c.value = text
    c.fill = fill(NAVY)
    c.font = Font(name="Calibri", size=size, bold=True, color="FFFFFF")
    c.alignment = center
    c.border = thin_border()


def _section_title(ws, row, code, title):
    ws.row_dimensions[row].height = 20
    ws.merge_cells(f"B{row}:E{row}")
    c = ws[f"B{row}"]
    c.value = f"Section {code}: {title}"
    from .utils import MED_BLUE
    c.fill = fill(MED_BLUE)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.alignment = left
    c.border = thin_border()
    return row + 1


def _label_row(ws, row, code, label, val, is_input=False, bold=False):
    cc = ws.cell(row=row, column=2, value=code)
    cc.fill = fill(ALT_ROW); cc.font = Font(name="Calibri", size=9, bold=True); cc.alignment = center; cc.border = thin_border()

    lc = ws.cell(row=row, column=3, value=label)
    lc.fill = fill(WHITE if not bold else "EBF5FB")
    lc.font = Font(name="Calibri", size=11, bold=bold)
    lc.alignment = left; lc.border = thin_border()

    vc = ws.cell(row=row, column=4, value=val if not is_input else (val if val is not None else None))
    vc.fill      = fill(INPUT_BG) if is_input else fill(WHITE if not bold else "EBF5FB")
    vc.font      = input_font() if is_input else Font(name="Calibri", size=11, bold=bold)
    vc.alignment = right_a; vc.border = thin_border()
    if val is not None and not is_input:
        pass  # formula will be set separately
    if is_input and val is not None:
        # percentage formatting for pct fields
        if 0 < val < 1:
            vc.number_format = "0.0%"
        else:
            vc.number_format = "$#,##0.00"

    uc = ws.cell(row=row, column=5, value="USD" if not is_input else "")
    uc.fill = fill(ALT_ROW); uc.font = Font(name="Calibri", size=9, color="888888"); uc.alignment = center; uc.border = thin_border()


def _make_bold_row(ws, row):
    for col in range(2, 6):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Calibri", size=11, bold=True)
        c.fill = fill(GRAY_SUB)


def _blank(ws, row, col, bg=WHITE):
    c = ws.cell(row=row, column=col)
    c.fill = fill(bg); c.border = thin_border()


def _set_formula(ws, row, formula, bold=False, pct=False, fmt=None):
    c = ws.cell(row=row, column=4, value=formula)
    c.font = Font(name="Calibri", size=11, bold=bold)
    c.alignment = right_a
    if pct:
        c.number_format = "0.0%"
    elif fmt:
        c.number_format = fmt
    else:
        c.number_format = "$#,##0.00"
