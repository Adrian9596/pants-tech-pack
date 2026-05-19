from .utils import (
    fill, bold_dark, bold_white, normal, thin_border, center, left,
    merge_title, style_header, set_col_widths, disable_gridlines,
    freeze, set_print_area, NAVY, MED_BLUE, WHITE, ALT_ROW
)
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter


def generate_sketch(wb, req, defaults):
    ws = wb.create_sheet("Technical Sketch")
    disable_gridlines(ws)

    # Column widths: A–Q left zone, R–AH right zone
    for col in range(1, 35):
        ws.column_dimensions[get_column_letter(col)].width = 9.5
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["R"].width = 2

    # ── Sheet title ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    ws.merge_cells("B2:Q2")
    _header(ws["B2"], f"{req.style_name} — Technical Sketch (Front)", 14)
    ws.merge_cells("S2:AH2")
    _header(ws["S2"], f"{req.style_name} — Technical Sketch (Back / Inner)", 14)

    # ── Section labels ────────────────────────────────────────────
    ws.row_dimensions[3].height = 16
    ws.merge_cells("B3:Q3")
    _subhdr(ws["B3"], "FRONT VIEW")
    ws.merge_cells("S3:AH3")
    _subhdr(ws["S3"], "BACK VIEW + INNER DETAILS")

    # ── Sketch placeholder (left / front) ─────────────────────────
    for r in range(4, 30):
        ws.row_dimensions[r].height = 20
    ws.merge_cells("B4:Q29")
    ph = ws["B4"]
    ph.value = (
        "[ FRONT FLAT SKETCH ]\n\n"
        "Upload sketch image to replace this placeholder.\n"
        "Front view — annotate with red arrows and bilingual EN/CN callout labels."
    )
    ph.fill      = fill("F0F4FA")
    ph.font      = Font(name="Calibri", size=11, italic=True, color="5A7FA8")
    ph.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ph.border    = Border(
        left=Side(style="medium", color="2E6DA4"),
        right=Side(style="medium", color="2E6DA4"),
        top=Side(style="medium", color="2E6DA4"),
        bottom=Side(style="medium", color="2E6DA4"),
    )

    # ── Sketch placeholder (right / back) ─────────────────────────
    ws.merge_cells("S4:AH25")
    ph2 = ws["S4"]
    ph2.value = (
        "[ BACK FLAT SKETCH ]\n\n"
        "Upload sketch image to replace this placeholder.\n"
        "Back view — annotate with red arrows and bilingual EN/CN callout labels."
    )
    ph2.fill      = fill("F0F4FA")
    ph2.font      = Font(name="Calibri", size=11, italic=True, color="5A7FA8")
    ph2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ph2.border    = Border(
        left=Side(style="medium", color="2E6DA4"),
        right=Side(style="medium", color="2E6DA4"),
        top=Side(style="medium", color="2E6DA4"),
        bottom=Side(style="medium", color="2E6DA4"),
    )

    # ── INNER section (right zone, below back sketch) ─────────────
    ws.row_dimensions[26].height = 20
    ws.merge_cells("S26:AH26")
    inner = ws["S26"]
    inner.value     = "INNER — 里布 / Interior Construction Details"
    inner.fill      = fill(MED_BLUE)
    inner.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    inner.alignment = center
    inner.border    = thin_border()

    for r in range(27, 42):
        ws.row_dimensions[r].height = 20
    ws.merge_cells("S27:AH41")
    inner_ph = ws["S27"]
    inner_ph.value = (
        "[ INNER CONSTRUCTION PHOTO / DETAIL ]\n\n"
        "Insert lining, inner structure, or cross-section detail photos here."
    )
    inner_ph.fill      = fill("F5F5F5")
    inner_ph.font      = Font(name="Calibri", size=10, italic=True, color="888888")
    inner_ph.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    inner_ph.border    = Border(
        left=Side(style="dashed", color="AAAAAA"),
        right=Side(style="dashed", color="AAAAAA"),
        top=Side(style="dashed", color="AAAAAA"),
        bottom=Side(style="dashed", color="AAAAAA"),
    )

    # ── Callout table (below front sketch) ────────────────────────
    callout_start = 31
    ws.row_dimensions[callout_start].height = 20
    merge_title(ws, callout_start, 2, 17, "CONSTRUCTION CALLOUTS", font_size=12)

    headers = ["#", "Callout Label (EN)", "Callout Label (CN)", "Detail / Specification"]
    col_map = {2: 1, 3: 6, 9: 6, 15: 3}  # col: span
    hrow = callout_start + 1
    ws.row_dimensions[hrow].height = 18

    # Header row for callout table
    for col, (hdr, span) in zip([2, 3, 9, 15],
                                 [("#", 1), ("Callout Label (EN)", 6),
                                  ("Callout Label (CN)", 6), ("Detail / Specification", 3)]):
        end = get_column_letter(col + span - 1)
        ws.merge_cells(f"{get_column_letter(col)}{hrow}:{end}{hrow}")
        c = ws.cell(row=hrow, column=col, value=hdr)
        style_header(c)

    # Callout rows
    callouts = defaults.get("callouts", [])
    for i in range(10):
        r = hrow + 1 + i
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        label_en = callouts[i][0] if i < len(callouts) else f"Callout {i+1}"
        detail   = callouts[i][1] if i < len(callouts) else ""

        # # cell
        nc = ws.cell(row=r, column=2, value=i + 1)
        nc.fill = fill(ALT_ROW if alt else WHITE); nc.font = bold_dark(); nc.alignment = center; nc.border = thin_border()

        # EN label
        ws.merge_cells(f"C{r}:H{r}")
        ec = ws.cell(row=r, column=3, value=label_en)
        ec.fill = fill(ALT_ROW if alt else WHITE); ec.font = normal(); ec.alignment = left; ec.border = thin_border()

        # CN placeholder
        ws.merge_cells(f"I{r}:N{r}")
        cc = ws.cell(row=r, column=9, value="")
        cc.fill = fill(ALT_ROW if alt else WHITE); cc.font = normal(); cc.alignment = left; cc.border = thin_border()

        # Detail
        ws.merge_cells(f"O{r}:Q{r}")
        dc = ws.cell(row=r, column=15, value=detail)
        dc.fill = fill(ALT_ROW if alt else WHITE); dc.font = normal(); dc.alignment = left; dc.border = thin_border()

    set_print_area(ws, "A1:AH55")
    freeze(ws, "A4")


def _header(cell, text, size=14):
    cell.value     = text
    cell.fill      = fill(NAVY)
    cell.font      = Font(name="Calibri", size=size, bold=True, color="FFFFFF")
    cell.alignment = center
    cell.border    = thin_border()


def _subhdr(cell, text):
    cell.value     = text
    cell.fill      = fill(MED_BLUE)
    cell.font      = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.alignment = center
    cell.border    = thin_border()
