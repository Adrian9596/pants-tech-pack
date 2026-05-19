from .utils import (
    fill, bold_dark, bold_white, normal, thin_border, center, left,
    merge_title, style_header, set_col_widths, disable_gridlines,
    freeze, set_print_area, NAVY, MED_BLUE, WHITE, ALT_ROW
)
from openpyxl.styles import Font


def generate_colorways(wb, req):
    ws = wb.create_sheet("Colorways & Lab Dips")
    disable_gridlines(ws)
    set_col_widths(ws, {1:18, 2:16, 3:16, 4:18, 5:16, 6:14, 7:14, 8:16, 9:22})

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    merge_title(ws, 2, 1, 9, f"COLORWAYS & LAB DIPS — {req.style_name}  |  {req.resolved_style_number()}", 14)

    # ── Colorway table ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 20
    merge_title(ws, 4, 1, 9, "COLORWAY SPECIFICATIONS", 12, bg=MED_BLUE)

    cw_headers = [
        "Colorway Name", "Primary Color", "Pantone #",
        "Secondary Color", "Pantone #", "Accent Color", "Pantone #",
        "Lab Dip Status", "Comments"
    ]
    ws.row_dimensions[5].height = 18
    for col, hdr in enumerate(cw_headers, 1):
        c = ws.cell(row=5, column=col, value=hdr)
        style_header(c)

    colorways = req.colorways if req.colorways else [
        type('CW', (), {'name':'CW1','primary_color':'','primary_pantone':'',
                        'secondary_color':'','secondary_pantone':'',
                        'accent_color':'','accent_pantone':''})(),
        type('CW', (), {'name':'CW2','primary_color':'','primary_pantone':'',
                        'secondary_color':'','secondary_pantone':'',
                        'accent_color':'','accent_pantone':''})(),
        type('CW', (), {'name':'CW3','primary_color':'','primary_pantone':'',
                        'secondary_color':'','secondary_pantone':'',
                        'accent_color':'','accent_pantone':''})(),
    ]

    for i, cw in enumerate(colorways):
        r = 6 + i
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        vals = [
            cw.name,
            cw.primary_color,   cw.primary_pantone,
            cw.secondary_color, cw.secondary_pantone,
            cw.accent_color,    cw.accent_pantone,
            "Pending",          ""
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(ALT_ROW if alt else WHITE)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = center if col in (3, 5, 7, 8) else left
            c.border    = thin_border()

    # Extra blank rows for more colorways
    for i in range(3):
        r = 6 + len(colorways) + i
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        for col in range(1, 10):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(ALT_ROW if alt else WHITE)
            c.border = thin_border()

    # ── Lab Dip Tracking table ─────────────────────────────────────
    lab_start = 6 + len(colorways) + 3 + 2
    ws.row_dimensions[lab_start].height = 20
    merge_title(ws, lab_start, 1, 9, "LAB DIP TRACKING", 12, bg=MED_BLUE)

    lab_hdrs = [
        "#", "Fabric / Trim", "Colorway", "Submission Date",
        "Lab Dip #", "Status", "Approval Date", "Approved By", "Comments"
    ]
    ws.row_dimensions[lab_start + 1].height = 18
    for col, hdr in enumerate(lab_hdrs, 1):
        c = ws.cell(row=lab_start + 1, column=col, value=hdr)
        style_header(c)

    for i in range(10):
        r = lab_start + 2 + i
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        for col in range(1, 10):
            c = ws.cell(row=r, column=col, value=i + 1 if col == 1 else "")
            c.fill      = fill(ALT_ROW if alt else WHITE)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = center if col in (1, 6, 7) else left
            c.border    = thin_border()

    last_row = lab_start + 12
    freeze(ws, "A5")
    set_print_area(ws, f"A1:I{last_row}")
