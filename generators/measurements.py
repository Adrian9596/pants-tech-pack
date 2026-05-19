from datetime import datetime
from .utils import (
    fill, bold_dark, bold_white, normal, thin_border, center, left,
    merge_title, style_header, set_col_widths, disable_gridlines, freeze,
    set_print_area, NAVY, MED_BLUE, WHITE, ALT_ROW
)
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from data import measurements_male, measurements_female


def generate_measurements(wb, req, defaults):
    ws = wb.create_sheet("Measurements")
    disable_gridlines(ws)

    chart_data = _load_chart(req.size_chart)

    # ── Header rows ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:Z2")
    c = ws["A2"]
    c.value     = f"SIZE SPECIFICATIONS — {req.style_name}  |  {req.resolved_style_number()}"
    c.fill      = fill(NAVY)
    c.font      = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.alignment = center
    c.border    = thin_border()

    ws.row_dimensions[3].height = 16
    meta_labels = [
        ("Garment Category", req.category or req.garment_type.replace("_", " ").title()),
        ("Season", req.season or "TBD"),
        ("Date", datetime.now().strftime("%Y-%m-%d")),
        ("Size Range", req.size_range),
        ("Measurement Standard", "Alvanon North America — inches"),
    ]
    for i, (k, v) in enumerate(meta_labels):
        col = i * 4 + 1
        kc = ws.cell(row=3, column=col, value=k)
        kc.font = Font(name="Calibri", bold=True, size=9); kc.alignment = left; kc.fill = fill(ALT_ROW); kc.border = thin_border()
        vc = ws.cell(row=3, column=col+1, value=v)
        vc.font = Font(name="Calibri", size=9); vc.alignment = left; vc.fill = fill(WHITE); vc.border = thin_border()

    ws.row_dimensions[4].height = 14
    ws.merge_cells("A4:Z4")
    tol = ws["A4"]
    tol.value     = "Default tolerance: ±1/4\" / ±0.5 cm unless stated per POM.  All measurements are ½ garment unless otherwise noted."
    tol.font      = Font(name="Calibri", size=9, italic=True, color="555555")
    tol.alignment = left
    tol.border    = thin_border()

    # ── Build size columns ─────────────────────────────────────────
    if chart_data:
        sizes = chart_data["sizes"]
    else:
        sizes = _parse_size_range(req.size_range)

    fixed_cols = 3   # #, Description, How to Measure
    col_pom   = 1
    col_desc  = 2
    col_how   = 3
    first_size_col = 4
    last_size_col  = first_size_col + len(sizes) - 1
    col_grade = last_size_col + 1
    col_tol   = col_grade + 1

    # Column widths
    set_col_widths(ws, {col_pom: 6, col_desc: 32, col_how: 36})
    for c in range(first_size_col, last_size_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 9
    ws.column_dimensions[get_column_letter(col_grade)].width = 14
    ws.column_dimensions[get_column_letter(col_tol)].width = 10

    # ── Column header row ─────────────────────────────────────────
    hrow = 5
    ws.row_dimensions[hrow].height = 20
    for col, hdr in zip(range(1, col_tol + 1),
                        ["#", "Measurement Description", "How to Measure (POM)"] +
                        sizes + ["Grading Increment", "Tolerance"]):
        c = ws.cell(row=hrow, column=col, value=hdr)
        c.fill      = fill(NAVY)
        c.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.alignment = center
        c.border    = thin_border()

    # ── Data rows ─────────────────────────────────────────────────
    pom_rows = chart_data["poms"] if chart_data else defaults.get("measurements", [])
    data_start = 6
    for i, row_data in enumerate(pom_rows):
        r = data_start + i
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        bg  = fill(ALT_ROW) if alt else fill(WHITE)

        if chart_data:
            pom_num, desc, tol_val, values = row_data
        else:
            if len(row_data) == 4:
                pom_num, desc, how, tol_val = row_data
                values = []
            else:
                pom_num, desc, tol_val = row_data[0], row_data[1], row_data[-1]
                how    = row_data[2] if len(row_data) > 3 else ""
                values = []

        # # column
        c = ws.cell(row=r, column=col_pom, value=pom_num)
        c.fill = bg; c.font = Font(name="Calibri", size=10, bold=True); c.alignment = center; c.border = thin_border()

        # Description
        if chart_data:
            how_text = ""
        else:
            how_text = how if len(row_data) > 3 else ""

        c = ws.cell(row=r, column=col_desc, value=desc)
        c.fill = bg; c.font = Font(name="Calibri", size=10); c.alignment = left; c.border = thin_border()

        # How to measure
        if not chart_data and len(row_data) > 3:
            how_val = row_data[2]
        else:
            how_val = ""
        c = ws.cell(row=r, column=col_how, value=how_val)
        c.fill = bg; c.font = Font(name="Calibri", size=9, italic=True, color="555555"); c.alignment = left; c.border = thin_border()

        # Size value cells
        for j, sz in enumerate(sizes):
            col = first_size_col + j
            val = values[j] if values and j < len(values) else None
            c = ws.cell(row=r, column=col, value=val)
            c.fill = bg; c.font = Font(name="Calibri", size=10); c.alignment = center; c.border = thin_border()
            if val is not None:
                c.number_format = '0.000'

        # Grading increment (blank — user fills)
        c = ws.cell(row=r, column=col_grade, value="")
        c.fill = fill("EBF5FB"); c.font = Font(name="Calibri", size=10); c.alignment = center; c.border = thin_border()

        # Tolerance
        c = ws.cell(row=r, column=col_tol, value=tol_val)
        c.fill = bg; c.font = Font(name="Calibri", size=10); c.alignment = center; c.border = thin_border()

    freeze(ws, "D6")
    set_print_area(ws, f"A1:{get_column_letter(col_tol)}{data_start + len(pom_rows) + 2}")


def _load_chart(size_chart: str):
    if not size_chart:
        return None
    key = size_chart.lower().replace(" ", "_").replace("-", "_")
    data = measurements_male.get_chart(key) or measurements_female.get_chart(key)
    return data


def _parse_size_range(size_range: str) -> list:
    presets = {
        "xs-xl":   ["XS", "S", "M", "L", "XL"],
        "xs-2xl":  ["XS", "S", "M", "L", "XL", "2XL"],
        "xs-4xl":  ["XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL"],
        "s-xl":    ["S", "M", "L", "XL"],
        "0-14":    ["0", "2", "4", "6", "8", "10", "12", "14"],
        "0-18":    ["0", "2", "4", "6", "8", "10", "12", "14", "16", "18"],
        "28-40":   ["28", "30", "32", "34", "36", "38", "40"],
    }
    key = size_range.lower().replace(" ", "")
    if key in presets:
        return presets[key]
    # split on / or ,
    import re
    parts = re.split(r"[/,]", size_range)
    if len(parts) > 1:
        return [p.strip() for p in parts]
    return ["XS", "S", "M", "L", "XL"]
