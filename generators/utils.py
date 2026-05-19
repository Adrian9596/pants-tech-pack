"""Shared openpyxl utilities for all sheet generators."""
from datetime import datetime
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Brand colors ──────────────────────────────────────────────────
NAVY       = "1A3A5C"
MED_BLUE   = "2E6DA4"
ORANGE     = "E8711A"
INPUT_BG   = "FFF3E0"   # light orange — user-input cells
FABRIC_BG  = "D6E4F0"   # light blue
TRIM_BG    = "D6F0D8"   # light green
LABEL_BG   = "FFF9C4"   # light yellow
ALT_ROW    = "F5F5F5"   # alternate body row
GRAY_SUB   = "D9D9D9"   # subtotal rows
WHITE      = "FFFFFF"
GREEN_OK   = "C6EFCE"   # margin traffic-light green
YELLOW_OK  = "FFEB9C"   # margin yellow
RED_WARN   = "FFC7CE"   # margin red
EXW_BG     = "DDEEFF"
FOB_BG     = "D6F0D8"


# ── Fonts ─────────────────────────────────────────────────────────
def bold_white(size=11):  return Font(name="Calibri", size=size, bold=True, color=WHITE)
def bold_navy(size=11):   return Font(name="Calibri", size=size, bold=True, color=NAVY)
def bold_dark(size=11):   return Font(name="Calibri", size=size, bold=True, color="000000")
def normal(size=11):      return Font(name="Calibri", size=size)
def input_font(size=11):  return Font(name="Calibri", size=size, color="1F4E79")


# ── Fills ─────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


# ── Borders ───────────────────────────────────────────────────────
_THIN  = Side(style="thin",   color="000000")
_THICK = Side(style="medium", color="000000")
_NONE  = Side(style=None)

def thin_border():
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

def thick_border():
    return Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)

def outer_thick(left=True, right=True, top=True, bottom=True):
    return Border(
        left=_THICK if left else _THIN,
        right=_THICK if right else _THIN,
        top=_THICK if top else _THIN,
        bottom=_THICK if bottom else _THIN,
    )


# ── Alignment ─────────────────────────────────────────────────────
center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
left    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
right_a = Alignment(horizontal="right",  vertical="center", wrap_text=True)


# ── Helpers ───────────────────────────────────────────────────────
def style_header(cell, size=11):
    """Navy fill, white bold text, centered."""
    cell.fill    = fill(NAVY)
    cell.font    = bold_white(size)
    cell.alignment = center
    cell.border  = thin_border()


def style_subheader(cell, size=11):
    cell.fill    = fill(MED_BLUE)
    cell.font    = bold_white(size)
    cell.alignment = center
    cell.border  = thin_border()


def style_input(cell):
    """Orange-fill input cell."""
    cell.fill      = fill(INPUT_BG)
    cell.font      = input_font()
    cell.alignment = left
    cell.border    = thin_border()


def style_data(cell, alt=False):
    cell.fill      = fill(ALT_ROW) if alt else fill(WHITE)
    cell.font      = normal()
    cell.alignment = left
    cell.border    = thin_border()


def style_subtotal(cell):
    cell.fill      = fill(GRAY_SUB)
    cell.font      = bold_dark()
    cell.alignment = left
    cell.border    = thin_border()


def set_col_widths(ws, widths: dict):
    """widths = {col_letter_or_int: width}"""
    for col, w in widths.items():
        ltr = get_column_letter(col) if isinstance(col, int) else col
        ws.column_dimensions[ltr].width = w


def merge_title(ws, row, start_col, end_col, text, font_size=14, bg=NAVY):
    start = get_column_letter(start_col)
    end   = get_column_letter(end_col)
    ws.merge_cells(f"{start}{row}:{end}{row}")
    cell = ws[f"{start}{row}"]
    cell.value     = text
    cell.fill      = fill(bg)
    cell.font      = bold_white(font_size)
    cell.alignment = center
    cell.border    = thin_border()
    ws.row_dimensions[row].height = 24


def money_fmt(cell):
    cell.number_format = '$#,##0.00'

def pct_fmt(cell):
    cell.number_format = '0.0%'


def auto_style_number() -> str:
    return f"STY-{datetime.now().year}-001"


def disable_gridlines(ws):
    ws.sheet_view.showGridLines = False


def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def set_print_area(ws, area: str):
    ws.print_area = area
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
