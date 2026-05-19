from .utils import (
    fill, bold_dark, bold_white, normal, input_font, thin_border, center, left, money_fmt,
    merge_title, style_header, set_col_widths, disable_gridlines, freeze, set_print_area,
    NAVY, WHITE, ALT_ROW, FABRIC_BG, TRIM_BG, LABEL_BG, INPUT_BG, GRAY_SUB
)
from openpyxl.styles import Font


_COLS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O"]
_HEADERS = [
    "#", "Item Type", "Item Description", "Supplier / Brand", "Color",
    "Pantone / Ref #", "Content", "Width / Size", "Placement",
    "Qty / Garment", "Unit", "Unit Cost (USD)", "Total Cost / Garment",
    "Lead Time (wks)", "Notes / Approval"
]
_WIDTHS = {1:4, 2:14, 3:30, 4:18, 5:12, 6:15, 7:18, 8:12, 9:18,
           10:10, 11:7, 12:14, 13:16, 14:12, 15:22}


def generate_bom(wb, req, defaults):
    ws = wb.create_sheet("BOM")
    disable_gridlines(ws)
    set_col_widths(ws, _WIDTHS)

    # ── Title ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    merge_title(ws, 2, 1, 15, f"BILL OF MATERIALS — {req.style_name}  |  {req.resolved_style_number()}", 14)

    # ── Column headers ─────────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    for col, hdr in enumerate(_HEADERS, start=1):
        c = ws.cell(row=3, column=col, value=hdr)
        style_header(c)

    # ── Data rows ─────────────────────────────────────────────────
    # Normalise: Pydantic models → plain dicts so item.get() works uniformly
    if req.bom_items:
        items = [i.model_dump() if hasattr(i, "model_dump") else dict(i) for i in req.bom_items]
    else:
        items = _build_defaults(defaults)
    data_start = 4
    for i, item in enumerate(items):
        r = data_start + i
        ws.row_dimensions[r].height = 16
        row_bg = _row_color(item["item_type"])
        alt    = i % 2 == 1

        def _cell(col, val, bg=None, fmt=None, is_input=False):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(bg or (INPUT_BG if is_input else row_bg))
            c.font      = input_font() if is_input else normal()
            c.alignment = center if col in (1, 10, 11, 12, 13, 14) else left
            c.border    = thin_border()
            if fmt:
                c.number_format = fmt
            return c

        _cell(1,  i + 1)
        _cell(2,  item.get("item_type", ""))
        _cell(3,  item.get("description", ""))
        _cell(4,  item.get("supplier", "TBD"))
        _cell(5,  item.get("color", ""))
        _cell(6,  item.get("pantone", ""))
        _cell(7,  item.get("content", ""))
        _cell(8,  item.get("width_size", ""))
        _cell(9,  item.get("placement", ""))
        _cell(10, item.get("qty", 1), fmt="0.000")
        _cell(11, item.get("unit", "yds"))

        # Unit Cost — orange input cell if not provided
        uc_val = item.get("unit_cost")
        uc     = _cell(12, uc_val, bg=INPUT_BG, fmt="$#,##0.00", is_input=True)

        # Total Cost — formula
        tc = ws.cell(row=r, column=13)
        tc.value        = f'=IF(L{r}="","",J{r}*L{r})'
        tc.fill         = fill(row_bg)
        tc.font         = normal()
        tc.alignment    = center
        tc.border       = thin_border()
        tc.number_format = "$#,##0.00"

        _cell(14, item.get("lead_time", 0), fmt="0")
        _cell(15, item.get("notes", ""))

    # ── Total Materials Cost row ───────────────────────────────────
    last_data = data_start + len(items) - 1
    total_row = last_data + 1
    ws.row_dimensions[total_row].height = 20

    for col in range(1, 16):
        c = ws.cell(row=total_row, column=col)
        c.fill   = fill(GRAY_SUB)
        c.font   = Font(name="Calibri", size=11, bold=True)
        c.border = thin_border()
        c.alignment = left

    ws.cell(row=total_row, column=3).value = "TOTAL MATERIALS COST"
    ws.cell(row=total_row, column=3).font  = Font(name="Calibri", size=11, bold=True)
    tc = ws.cell(row=total_row, column=13)
    tc.value          = f"=SUM(M{data_start}:M{last_data})"
    tc.number_format  = "$#,##0.00"
    tc.alignment      = center

    freeze(ws, "A4")
    set_print_area(ws, f"A1:O{total_row + 2}")


def _row_color(item_type: str) -> str:
    t = item_type.lower()
    if "fabric"   in t: return FABRIC_BG
    if "trim"     in t: return TRIM_BG
    return LABEL_BG


def _build_defaults(defaults: dict):
    return [
        {
            "item_type":   item.get("item_type", "Fabric"),
            "description": item.get("description", ""),
            "supplier":    item.get("supplier", "TBD"),
            "color":       item.get("color", ""),
            "pantone":     item.get("pantone", ""),
            "content":     item.get("content", ""),
            "width_size":  item.get("width_size", ""),
            "placement":   item.get("placement", ""),
            "qty":         item.get("qty", 1),
            "unit":        item.get("unit", "yds"),
            "unit_cost":   item.get("unit_cost"),
            "lead_time":   item.get("lead_time", 0),
            "notes":       item.get("notes", ""),
        }
        for item in defaults.get("bom", [])
    ]
