from datetime import datetime
from openpyxl.styles import Alignment
from .utils import (
    fill, bold_white, bold_navy, bold_dark, normal, input_font,
    thin_border, center, left, merge_title, style_header, set_col_widths,
    disable_gridlines, freeze, set_print_area,
    NAVY, MED_BLUE, WHITE, ALT_ROW, ORANGE
)
from openpyxl.utils import get_column_letter


def generate_cover(wb, req):
    ws = wb.create_sheet("Cover Page")
    disable_gridlines(ws)

    set_col_widths(ws, {1: 5, 2: 25, 3: 35, 4: 20, 5: 25, 6: 5})
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 14

    # ── Big brand header ──────────────────────────────────────────
    ws.merge_cells("B2:E2")
    c = ws["B2"]
    c.value     = req.brand_name
    c.fill      = fill(NAVY)
    c.font      = bold_white(22)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()

    # Sub-header bar
    ws.row_dimensions[4].height = 20
    ws.merge_cells("B4:E4")
    sub = ws["B4"]
    sub.value     = "TECHNICAL PACKAGE"
    sub.fill      = fill(MED_BLUE)
    sub.font      = bold_white(13)
    sub.alignment = center
    sub.border    = thin_border()

    # ── Metadata table ────────────────────────────────────────────
    rows = [
        ("Style Name",     req.style_name),
        ("Style Number",   req.resolved_style_number()),
        ("Season / Year",  req.season or "TBD"),
        ("Category",       req.category or req.garment_type.replace("_", " ").title()),
        ("Designer",       req.designer),
        ("Date Created",   datetime.now().strftime("%Y-%m-%d")),
        ("Revision #",     "Rev 1"),
        ("Factory",        req.factory_name),
        ("Target FOB",     req.target_fob),
        ("Fabric Content", req.fabric_content or "TBD"),
        ("Size Range",     req.size_range),
    ]

    start_row = 6
    for i, (label, value) in enumerate(rows):
        r = start_row + i
        ws.row_dimensions[r].height = 18
        alt = i % 2 == 1

        # Label cell
        lc = ws.cell(row=r, column=2, value=label)
        lc.fill      = fill(ALT_ROW) if alt else fill(WHITE)
        lc.font      = bold_dark()
        lc.alignment = left
        lc.border    = thin_border()

        # Value cell (spans C–E)
        ws.merge_cells(f"C{r}:E{r}")
        vc = ws.cell(row=r, column=3, value=value)
        vc.fill      = fill(ALT_ROW) if alt else fill(WHITE)
        vc.font      = normal()
        vc.alignment = left
        vc.border    = thin_border()

    # ── Colorways row ─────────────────────────────────────────────
    cw_row = start_row + len(rows)
    ws.row_dimensions[cw_row].height = 18
    lc = ws.cell(row=cw_row, column=2, value="Colorways")
    lc.fill = fill(WHITE); lc.font = bold_dark(); lc.alignment = left; lc.border = thin_border()
    ws.merge_cells(f"C{cw_row}:E{cw_row}")
    cws = " / ".join([c.name for c in req.colorways]) if req.colorways else "CW1 / CW2 / CW3"
    vc = ws.cell(row=cw_row, column=3, value=cws)
    vc.fill = fill(WHITE); vc.font = normal(); vc.alignment = left; vc.border = thin_border()

    # ── Revision history ──────────────────────────────────────────
    rev_start = cw_row + 2
    ws.row_dimensions[rev_start].height = 20
    merge_title(ws, rev_start, 2, 5, "REVISION HISTORY", font_size=12)

    for col, hdr in zip(range(2, 6), ["Rev #", "Date", "Description", "Author"]):
        c = ws.cell(row=rev_start + 1, column=col, value=hdr)
        style_header(c)
        ws.row_dimensions[rev_start + 1].height = 18

    rev_data = [
        ("Rev 1", datetime.now().strftime("%Y-%m-%d"), "Initial release", req.designer),
    ]
    for j, (rv, dt, desc, auth) in enumerate(rev_data):
        r = rev_start + 2 + j
        ws.row_dimensions[r].height = 16
        for col, val in zip(range(2, 6), [rv, dt, desc, auth]):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(ALT_ROW if j % 2 else WHITE)
            c.font = normal()
            c.alignment = left
            c.border = thin_border()

    # blank rows for future revisions
    for j in range(3):
        r = rev_start + 2 + len(rev_data) + j
        ws.row_dimensions[r].height = 16
        for col in range(2, 6):
            c = ws.cell(row=r, column=col, value="")
            c.fill = fill(ALT_ROW if j % 2 else WHITE)
            c.border = thin_border()

    set_print_area(ws, "A1:F50")
    freeze(ws, "A5")
