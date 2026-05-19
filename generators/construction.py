from .utils import (
    fill, bold_dark, bold_white, normal, thin_border, center, left,
    merge_title, style_header, set_col_widths, disable_gridlines,
    freeze, set_print_area, NAVY, MED_BLUE, WHITE, ALT_ROW, GRAY_SUB
)
from openpyxl.styles import Font, Alignment


_OUTERWEAR_STITCH_SPECS = [
    ("Side Seams",       "4-thread overlock + seam tape", "514 Overlock", "12 SPI", "Coats T40", "Apply seam tape; FG-OT-03"),
    ("Shoulder Seams",   "3-thread overlock + tape",      "504 Overlock", "12 SPI", "Coats T40", "Reinforce with seam tape"),
    ("Sleeve Set",       "4-thread overlock",             "514 Overlock", "12 SPI", "Match shell", ""),
    ("CF Zipper",        "Plain seam, topstitch 1/4\"",  "301 Lockstitch","12 SPI", "Match shell", "Bartack zipper top/bottom"),
    ("Chest Pocket",     "Topstitch 1/8\"",              "301 Lockstitch","12 SPI", "Match shell", "Bartack all corners"),
    ("Hem",              "Turned & topstitched 1\"",      "301 Lockstitch","12 SPI", "Match shell", "Drawcord channel"),
    ("Cuff Rib",         "3-thread overlock",             "504 Overlock", "12 SPI", "Match rib",   "Velcro tab after"),
    ("Lining",           "Hand slipstitch or blind hem",  "301 Lockstitch","12 SPI", "Match lining",""),
    ("Hood",             "Plain seam + overlock",         "301 + 504",    "12 SPI", "Match shell", "Topstitch 1/4\" at seam"),
]

_OUTERWEAR_QC = [
    ("Fabric Content",    "±3% of declared content — AATCC 20",       "1.5"),
    ("Fabric Weight",     "±5% of declared GSM — ASTM D3776",          "1.5"),
    ("Seam Strength",     "Min 22 lbf woven / 18 lbf knit — ASTM D1683","2.5"),
    ("Stitch Quality",    "No skipped stitches, min 10–12 SPI",         "1.5"),
    ("Zipper Operation",  "No snags; pull force ≤ 30N — ASTM D2062",   "1.5"),
    ("DWR / Water Repel", "Before wash ≥90, after 3 washes ≥80 — AATCC 22","2.5"),
    ("Pilling Resistance","≥3–4 — ASTM D3512",                         "2.5"),
    ("Dimensional Stability","Woven ≤3%; Knit ≤5% — AATCC 135",       "2.5"),
    ("Color Fastness",    "Change ≥4.0; Crocking dry ≥3.5 — AATCC 61/8","2.5"),
    ("Label Placement",   "Brand/size/care/fill labels correctly placed","1.5"),
    ("Appearance",        "Clean finish, no puckering, uniform seams",  "1.5"),
    ("Measurement",       "All POMs within stated tolerances",          "2.5"),
]


def generate_construction(wb, req, defaults):
    ws = wb.create_sheet("Construction Details")
    disable_gridlines(ws)
    set_col_widths(ws, {1:4, 2:22, 3:22, 4:18, 5:10, 6:14, 7:28})

    # ── Title ──────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 22
    merge_title(ws, 2, 1, 7, f"CONSTRUCTION DETAILS — {req.style_name}  |  {req.resolved_style_number()}", 14)

    is_outerwear = req.garment_type == "outerwear"
    cur_row = 4

    # ── Section A: Seam & Stitch ───────────────────────────────────
    ws.row_dimensions[cur_row].height = 20
    merge_title(ws, cur_row, 1, 7, "SECTION A — SEAM & STITCH SPECIFICATIONS", 12, bg=MED_BLUE)
    cur_row += 1

    a_hdrs = ["#", "Seam Location", "Seam Type", "Stitch Type", "SPI", "Thread", "Notes"]
    ws.row_dimensions[cur_row].height = 18
    for col, hdr in enumerate(a_hdrs, 1):
        c = ws.cell(row=cur_row, column=col, value=hdr)
        style_header(c)
    cur_row += 1

    seam_rows = _OUTERWEAR_STITCH_SPECS if is_outerwear else defaults.get("seam_specs", [])
    for i, row_data in enumerate(seam_rows):
        r = cur_row
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val if col > 1 else i + 1)
            c.fill      = fill(ALT_ROW if alt else WHITE)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = center if col in (1, 4, 5) else left
            c.border    = thin_border()
        cur_row += 1
    cur_row += 1

    # ── Section B: Finishing Instructions ─────────────────────────
    ws.row_dimensions[cur_row].height = 20
    merge_title(ws, cur_row, 1, 7, "SECTION B — FINISHING INSTRUCTIONS", 12, bg=MED_BLUE)
    cur_row += 1

    finishing = defaults.get("finishing", [])
    if is_outerwear:
        finishing = [
            "Apply seam sealing tape to all major structural seams (sides, shoulders)",
            "Topstitch all external seams 1/4\" for visual definition",
            "Bartack zipper ends, pocket corners, drawcord exits",
            "Test all zippers for smooth operation before shipping",
            "Insert drawcords; tie barrel-lock stops at both ends",
            "Press with low heat — no direct iron on nylon/shell",
            "Brand + size labels: CB neck, sewn all 4 sides",
            "Fill weight label + care label: left side seam",
            "Hang tag: front zipper pull with clear elastic",
            "Pack flat with tissue, poly bag, pressure-reseal",
        ]

    for i, inst in enumerate(finishing):
        ws.row_dimensions[cur_row].height = 16
        ws.merge_cells(f"A{cur_row}:G{cur_row}")
        c = ws.cell(row=cur_row, column=1, value=f"• {inst}")
        c.fill      = fill(ALT_ROW if i % 2 else WHITE)
        c.font      = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border    = thin_border()
        cur_row += 1
    cur_row += 1

    # ── Section C: QC Checkpoints ──────────────────────────────────
    ws.row_dimensions[cur_row].height = 20
    merge_title(ws, cur_row, 1, 7, "SECTION C — QUALITY CHECKPOINTS", 12, bg=MED_BLUE)
    cur_row += 1

    qc_hdrs = ["#", "Inspection Point", "Accept Criteria", "AQL Level", "", "", ""]
    ws.row_dimensions[cur_row].height = 18
    for col, hdr in enumerate(qc_hdrs, 1):
        c = ws.cell(row=cur_row, column=col, value=hdr)
        style_header(c)
    cur_row += 1

    qc_rows = _OUTERWEAR_QC if is_outerwear else _generic_qc()
    for i, (point, criteria, aql) in enumerate(qc_rows):
        r = cur_row
        ws.row_dimensions[r].height = 16
        alt = i % 2 == 1
        for col, val in enumerate([i + 1, point, criteria, aql, "", "", ""], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill      = fill(ALT_ROW if alt else WHITE)
            c.font      = Font(name="Calibri", size=10)
            c.alignment = center if col in (1, 4) else left
            c.border    = thin_border()
        cur_row += 1

    freeze(ws, "A4")
    set_print_area(ws, f"A1:G{cur_row + 2}")


def _generic_qc():
    return [
        ("Fabric Content",    "Matches declared content label",            "1.5"),
        ("Seam Strength",     "No seam failure at standard pull test",      "2.5"),
        ("Stitch Quality",    "No skipped stitches, even SPI throughout",   "1.5"),
        ("Measurement",       "All POMs within stated tolerances",          "2.5"),
        ("Label Placement",   "All labels correctly positioned and sewn",   "1.5"),
        ("Zipper / Closure",  "Smooth operation, no snags",                 "1.5"),
        ("Color / Shade",     "Matches approved lab dip or standard",       "2.5"),
        ("Appearance",        "No puckering, broken stitches, or soil",     "1.5"),
    ]
