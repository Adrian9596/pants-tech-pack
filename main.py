import io
from fastapi import FastAPI, HTTPException
from fastapi.responses import Response, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook

from models import TechPackRequest
from data.garment_defaults import get_defaults, list_garment_types
from generators.cover_page import generate_cover
from generators.technical_sketch import generate_sketch
from generators.bom import generate_bom
from generators.costing import generate_costing
from generators.measurements import generate_measurements
from generators.construction import generate_construction
from generators.colorways import generate_colorways

app = FastAPI(title="Pants Tech Pack")

app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
def index():
    return FileResponse("static/index.html")


@app.get("/api/garment-types")
def garment_types():
    return list_garment_types()


@app.get("/api/defaults/{garment_type}")
def garment_defaults(garment_type: str):
    d = get_defaults(garment_type)
    return {
        "bom": d.get("bom", []),
        "callouts": d.get("callouts", []),
    }


@app.post("/api/generate")
def generate(req: TechPackRequest):
    try:
        defaults = get_defaults(req.garment_type)
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        if req.should_generate("cover"):
            generate_cover(wb, req)

        if req.should_generate("sketch"):
            generate_sketch(wb, req, defaults)

        if req.should_generate("bom"):
            generate_bom(wb, req, defaults)

        if req.should_generate("costing"):
            generate_costing(wb, req)

        if req.should_generate("measurements"):
            generate_measurements(wb, req, defaults)

        if req.should_generate("construction"):
            generate_construction(wb, req, defaults)

        if req.should_generate("colorways"):
            generate_colorways(wb, req)

        if not wb.sheetnames:
            raise HTTPException(status_code=422, detail="No sheets selected")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = req.style_name.replace(" ", "_").replace("/", "-") or "StyleName"
        filename  = f"{safe_name}_techpack.xlsx"

        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        raise  # pass through 422s and other intentional HTTP errors unchanged
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Generation failed: {type(e).__name__}: {e}",
        )
